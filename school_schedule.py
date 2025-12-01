import mysql.connector
from mysql.connector import Error
import pulp
import tkinter as tk
from tkinter import messagebox, ttk
from datetime import datetime
import _tkinter  # Para capturar TclError no try-except
import os
import logging
import argparse
import sys

# Configurações de conexão com o banco de dados MySQL
DB_HOST = os.environ.get('DB_HOST', 'localhost')
DB_USER = os.environ.get('DB_USER', 'root')
DB_PASSWORD = os.environ.get('DB_PASSWORD', '221203Ma')
DB_NAME = os.environ.get('DB_NAME', 'sistema_escolar')

# Função para criar a conexão com o banco de dados
def create_connection():
    connection = None
    try:
        connection = mysql.connector.connect(
            host=DB_HOST,
            user=DB_USER,
            passwd=DB_PASSWORD,
            database=DB_NAME
        )
        logging.info("Conexão com MySQL DB bem-sucedida")
    except Error as e:
        logging.error(f"Erro ao conectar ao MySQL DB: '{e}'")
    return connection

# Função para criar as tabelas no banco de dados
def create_tables(connection):
    cursor = connection.cursor()
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS professores (
        id INT AUTO_INCREMENT PRIMARY KEY,
        nome VARCHAR(100) NOT NULL,
        disponibilidade TEXT,
        preferencias TEXT
    )
    """)
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS materias (
        id INT AUTO_INCREMENT PRIMARY KEY,
        nome VARCHAR(100) NOT NULL,
        carga_horaria INT NOT NULL
    )
    """)
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS turmas (
        id INT AUTO_INCREMENT PRIMARY KEY,
        nome VARCHAR(50) NOT NULL,
        ano INT NOT NULL
    )
    """)
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS cronogramas (
        id INT AUTO_INCREMENT PRIMARY KEY,
        professor_id INT,
        materia_id INT,
        turma_id INT,
        dia_semana VARCHAR(20),
        horario_inicio TIME,
        horario_fim TIME,
        FOREIGN KEY (professor_id) REFERENCES professores(id),
        FOREIGN KEY (materia_id) REFERENCES materias(id),
        FOREIGN KEY (turma_id) REFERENCES turmas(id)
    )
    """)
    
    connection.commit()
    print("Tabelas criadas com sucesso")

# Função para mostrar mensagem (compatível com CI/headless)
def show_message(title, message, type="info"):
    try:
        import tkinter.messagebox as mb
        if type == "info":
            mb.showinfo(title, message)
        elif type == "error":
            mb.showerror(title, message)
    except (ImportError, _tkinter.TclError):
        logging.info(f"{title}: {message} (Modo headless)")

# Função para otimizar o cronograma usando PuLP
def optimize_schedule(connection):
    cursor = connection.cursor()
    
    # Verificar se há dados suficientes
    cursor.execute("SELECT COUNT(*) FROM professores")
    if cursor.fetchone()[0] == 0:
        show_message("Erro", "Nenhum professor cadastrado!", "error")
        return
    cursor.execute("SELECT COUNT(*) FROM materias")
    if cursor.fetchone()[0] == 0:
        show_message("Erro", "Nenhuma matéria cadastrada!", "error")
        return
    cursor.execute("SELECT COUNT(*) FROM turmas")
    if cursor.fetchone()[0] == 0:
        show_message("Erro", "Nenhuma turma cadastrada!", "error")
        return
    
    cursor.execute("SELECT id, nome, disponibilidade, preferencias FROM professores")
    professores = cursor.fetchall()
    
    cursor.execute("SELECT id, nome, carga_horaria FROM materias")
    materias = cursor.fetchall()
    
    cursor.execute("SELECT id, nome FROM turmas")
    turmas = cursor.fetchall()
    
    dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta']
    
    horarios = [
        ('08:00:00', '09:00:00'),
        ('09:00:00', '10:00:00'),
        ('10:00:00', '11:00:00'),
        ('11:00:00', '12:00:00')
    ]
    
    # Basic feasibility check: ensure total required slots <= available slots
    total_slots = len(dias_semana) * len(horarios) * len(turmas)
    total_required = sum(m[2] for m in materias)
    if total_required > total_slots:
        show_message("Erro", f"Carga horária total ({total_required}) excede slots disponíveis ({total_slots})", "error")
        return False

    prob = pulp.LpProblem("Agendamento_Escolar", pulp.LpMinimize)

    x = pulp.LpVariable.dicts("assign", 
                              ((p[0], m[0], t[0], d, s) 
                               for p in professores 
                               for m in materias 
                               for t in turmas 
                               for d in dias_semana 
                               for s in range(len(horarios))),
                              cat='Binary')

    # Penalidade para descumprimento de preferência (higher penalty -> stronger preference)
    penalty = 5
    obj = []
    for key in x:
        p_id, m_id, t_id, d, s = key
        prof = next((prof for prof in professores if prof[0] == p_id), None)
        penal = 0
        if prof and prof[3]:
            pref = prof[3]
            if ':' in pref:
                try:
                    pref_mat, pref_dia = pref.split(':', 1)
                except Exception:
                    pref_mat, pref_dia = '', ''
                mat_nome = next((mat[1] for mat in materias if mat[0] == m_id), "")
                # penalize mismatch; if both match prefer strongly
                if pref_mat and mat_nome != pref_mat:
                    penal += penalty
                if pref_dia and d != pref_dia:
                    penal += penalty
        # base cost 1 for an assignment; add penal when preference not matched
        obj.append(x[key] * (1 + penal))
    prob += pulp.lpSum(obj)

    for m_id, _, carga in materias:
        for t_id, _ in turmas:
            prob += pulp.lpSum(x[p[0], m_id, t_id, d, s] 
                               for p in professores 
                               for d in dias_semana 
                               for s in range(len(horarios))) == carga

    for p in professores:
        p_id = p[0]
        for d in dias_semana:
            for s in range(len(horarios)):
                prob += pulp.lpSum(x[p_id, m[0], t[0], d, s] 
                                   for m in materias 
                                   for t in turmas) <= 1

    # Garantir que para cada turma, dia e horário haja no máximo UMA atribuição (evita dois professores no mesmo dia/horário para mesma turma)
    for t in turmas:
        t_id = t[0]
        for d in dias_semana:
            for s in range(len(horarios)):
                prob += pulp.lpSum(x[p[0], m[0], t_id, d, s]
                                   for p in professores
                                   for m in materias) <= 1

    for p in professores:
        p_id = p[0]
        disp = p[2].split(',') if p[2] else []
        for d in dias_semana:
            if d not in disp:
                for m in materias:
                    for t in turmas:
                        for s in range(len(horarios)):
                            prob += x[p_id, m[0], t[0], d, s] == 0

    prob.solve(pulp.PULP_CBC_CMD(msg=0))

    if pulp.LpStatus[prob.status] != 'Optimal':
        logging.error(f"Solver status: {pulp.LpStatus[prob.status]}")
        show_message("Erro", "Não foi possível gerar um cronograma. Verifique os dados inseridos!", "error")
        return False

    cursor.execute("DELETE FROM cronogramas")
    for var in x:
        if pulp.value(x[var]) == 1:
            p_id, m_id, t_id, d, s = var
            inicio, fim = horarios[s]
            cursor.execute("""
            INSERT INTO cronogramas (professor_id, materia_id, turma_id, dia_semana, horario_inicio, horario_fim)
            VALUES (%s, %s, %s, %s, %s, %s)
            """, (p_id, m_id, t_id, d, inicio, fim))
    
    connection.commit()
    show_message("Sucesso", "Cronograma gerado com sucesso!", "info")
    logging.info("Cronograma gerado e salvo no banco")
    # Verificação pós-solução: detectar conflitos residuais (duas atribuições no mesmo slot)
    try:
        # Por turma/dia/horário
        cursor.execute("""
        SELECT turma_id, dia_semana, horario_inicio, COUNT(*) as cnt
        FROM cronogramas
        GROUP BY turma_id, dia_semana, horario_inicio
        HAVING cnt > 1
        """)
        conflicts = cursor.fetchall()
        if conflicts:
            msg = "Conflitos detectados por turma/dia/horário:\n"
            for tid, dia, inicio, cnt in conflicts:
                msg += f"- Turma {tid} {dia} {inicio} -> {cnt} atribuições\n"
            logging.error(msg)
            show_message("Conflitos", msg, "error")
            return False

        # Por professor/dia/horário (double-booking)
        cursor.execute("""
        SELECT professor_id, dia_semana, horario_inicio, COUNT(*) as cnt
        FROM cronogramas
        GROUP BY professor_id, dia_semana, horario_inicio
        HAVING cnt > 1
        """)
        prof_conf = cursor.fetchall()
        if prof_conf:
            msg = "Conflitos de professor (double-booking):\n"
            for pid, dia, inicio, cnt in prof_conf:
                msg += f"- Professor {pid} {dia} {inicio} -> {cnt} atribuições\n"
            logging.error(msg)
            show_message("Conflitos", msg, "error")
            return False
    except Exception as e:
        logging.error(f"Erro ao verificar conflitos: {e}")

    return True

# GUI para o diretor inserir dados
class SchoolApp:
    def __init__(self, root, connection):
        self.root = root
        self.conn = connection
        self.root.title("Sistema de Agendamento Escolar")
        self.root.geometry("900x700")
        self.root.configure(bg="#f0f4f8")

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TNotebook', background="#e3eaf2", borderwidth=0)
        style.configure('TFrame', background="#f0f4f8")
        style.configure('TLabel', background="#f0f4f8", font=("Segoe UI", 12))
        style.configure('TButton', font=("Segoe UI", 11, "bold"), padding=6)
        style.configure('TEntry', font=("Segoe UI", 11))
        style.configure('TCombobox', font=("Segoe UI", 11))

        self.header = tk.Label(root, text="Sistema de Agendamento Escolar", font=("Segoe UI", 20, "bold"), bg="#f0f4f8", fg="#2a3b4c")
        self.header.pack(pady=(15, 5))

        self.notebook = ttk.Notebook(root, style='TNotebook')
        self.notebook.pack(pady=10, padx=10, fill="both", expand=True)

        # Aba Professores
        self.prof_frame = ttk.Frame(self.notebook, style='TFrame')
        self.notebook.add(self.prof_frame, text="Professores")
        self.setup_prof_frame()

        # Aba Matérias
        self.mat_frame = ttk.Frame(self.notebook, style='TFrame')
        self.notebook.add(self.mat_frame, text="Matérias")
        self.setup_mat_frame()

        # Aba Turmas
        self.tur_frame = ttk.Frame(self.notebook, style='TFrame')
        self.notebook.add(self.tur_frame, text="Turmas")
        self.setup_tur_frame()

        # Aba Dados Cadastrados
        self.data_frame = ttk.Frame(self.notebook, style='TFrame')
        self.notebook.add(self.data_frame, text="Dados Cadastrados")
        self.setup_data_frame()

        # Aba Cronograma (tabela)
        self.schedule_frame = ttk.Frame(self.notebook, style='TFrame')
        self.notebook.add(self.schedule_frame, text="Cronograma")
        self.setup_schedule_frame()

        # Botão para gerar cronograma
        self.generate_btn = ttk.Button(root, text="Gerar Cronograma Otimizado", command=self.generate, style='TButton')
        self.generate_btn.pack(pady=10)

        # Área para exibir cronograma
        self.text_area = tk.Text(root, height=18, width=100, font=("Consolas", 11), bg="#f8fafc", fg="#222", borderwidth=2, relief="groove")
        self.text_area.pack(pady=10, padx=20)

    def setup_prof_frame(self):
        dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta']
        ttk.Label(self.prof_frame, text="Nome do Professor:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.prof_nome = tk.Entry(self.prof_frame, font=("Segoe UI", 11))
        self.prof_nome.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        ttk.Label(self.prof_frame, text="Disponibilidade (selecione):").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.prof_disp_vars = {}
        dias_frame = ttk.Frame(self.prof_frame)
        dias_frame.grid(row=1, column=1, padx=0, pady=0, sticky="w")
        for i, dia in enumerate(dias_semana):
            var = tk.BooleanVar()
            cb = ttk.Checkbutton(dias_frame, text=dia, variable=var)
            cb.pack(side="left", padx=2)
            self.prof_disp_vars[dia] = var

        ttk.Label(self.prof_frame, text="Preferência (Matéria):").grid(row=2, column=0, padx=10, pady=10, sticky="w")
        self.prof_pref_mat = tk.Entry(self.prof_frame, font=("Segoe UI", 11))
        self.prof_pref_mat.grid(row=2, column=1, padx=10, pady=10, sticky="ew")

        ttk.Label(self.prof_frame, text="Preferência (Dia da Semana):").grid(row=3, column=0, padx=10, pady=10, sticky="w")
        self.prof_pref_dia = ttk.Combobox(self.prof_frame, values=dias_semana, state="readonly", font=("Segoe UI", 11))
        self.prof_pref_dia.grid(row=3, column=1, padx=10, pady=10, sticky="ew")

        ttk.Button(self.prof_frame, text="Adicionar Professor", command=self.add_prof).grid(row=4, column=0, columnspan=2, pady=15)
        ttk.Button(self.prof_frame, text="Limpar Campos", command=self.clear_prof).grid(row=5, column=0, columnspan=2, pady=5)

        # Lista de professores cadastrados
        cols = ("id", "nome", "disponibilidade", "preferencias")
        self.prof_tree = ttk.Treeview(self.prof_frame, columns=cols, show='headings', height=6)
        for c in cols:
            self.prof_tree.heading(c, text=c.capitalize())
        self.prof_tree.grid(row=6, column=0, columnspan=6, padx=10, pady=10, sticky='nsew')

        btn_frame = ttk.Frame(self.prof_frame)
        btn_frame.grid(row=7, column=0, columnspan=6, pady=5)
        ttk.Button(btn_frame, text="Editar Selecionado", command=self.edit_prof).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Excluir Selecionado", command=self.delete_prof).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Atualizar Lista", command=self.refresh_prof_list).pack(side='left', padx=5)

        # Ajuste grid weights
        self.prof_frame.columnconfigure(1, weight=1)
        self.refresh_prof_list()

    def setup_mat_frame(self):
        ttk.Label(self.mat_frame, text="Nome da Matéria:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.mat_nome = tk.Entry(self.mat_frame, font=("Segoe UI", 11))
        self.mat_nome.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        ttk.Label(self.mat_frame, text="Carga Horária (número):").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.mat_carga = tk.Entry(self.mat_frame, font=("Segoe UI", 11))
        self.mat_carga.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

        ttk.Button(self.mat_frame, text="Adicionar Matéria", command=self.add_mat).grid(row=2, column=0, columnspan=2, pady=15)
        ttk.Button(self.mat_frame, text="Limpar Campos", command=self.clear_mat).grid(row=3, column=0, columnspan=2, pady=5)

        cols = ("id", "nome", "carga_horaria")
        self.mat_tree = ttk.Treeview(self.mat_frame, columns=cols, show='headings', height=6)
        for c in cols:
            self.mat_tree.heading(c, text=c.capitalize())
        self.mat_tree.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky='nsew')

        btn_frame = ttk.Frame(self.mat_frame)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=5)
        ttk.Button(btn_frame, text="Editar Selecionado", command=self.edit_mat).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Excluir Selecionado", command=self.delete_mat).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Atualizar Lista", command=self.refresh_mat_list).pack(side='left', padx=5)

        self.mat_frame.columnconfigure(1, weight=1)
        self.refresh_mat_list()

    def setup_tur_frame(self):
        ttk.Label(self.tur_frame, text="Nome da Turma:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.tur_nome = tk.Entry(self.tur_frame, font=("Segoe UI", 11))
        self.tur_nome.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        ttk.Label(self.tur_frame, text="Ano (número):").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.tur_ano = tk.Entry(self.tur_frame, font=("Segoe UI", 11))
        self.tur_ano.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

        ttk.Button(self.tur_frame, text="Adicionar Turma", command=self.add_tur).grid(row=2, column=0, columnspan=2, pady=15)
        ttk.Button(self.tur_frame, text="Limpar Campos", command=self.clear_tur).grid(row=3, column=0, columnspan=2, pady=5)

        cols = ("id", "nome", "ano")
        self.tur_tree = ttk.Treeview(self.tur_frame, columns=cols, show='headings', height=6)
        for c in cols:
            self.tur_tree.heading(c, text=c.capitalize())
        self.tur_tree.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky='nsew')

        btn_frame = ttk.Frame(self.tur_frame)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=5)
        ttk.Button(btn_frame, text="Editar Selecionado", command=self.edit_tur).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Excluir Selecionado", command=self.delete_tur).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Atualizar Lista", command=self.refresh_tur_list).pack(side='left', padx=5)

        self.tur_frame.columnconfigure(1, weight=1)
        self.refresh_tur_list()

    def setup_data_frame(self):
        self.data_text = tk.Text(self.data_frame, height=15, width=80, font=("Consolas", 11), bg="#f8fafc", fg="#222", borderwidth=2, relief="groove")
        self.data_text.pack(pady=10, padx=10)
        ttk.Button(self.data_frame, text="Atualizar Lista", command=self.list_data).pack(pady=5)

    def setup_schedule_frame(self):
        # Treeview para exibir o cronograma como tabela
        cols = ("professor", "materia", "turma", "dia", "inicio", "fim")
        self.schedule_tree = ttk.Treeview(self.schedule_frame, columns=cols, show='headings')
        for c in cols:
            hdr = c.capitalize()
            self.schedule_tree.heading(c, text=hdr)
            self.schedule_tree.column(c, width=120, anchor='center')
        self.schedule_tree.pack(fill='both', expand=True, padx=8, pady=8)

        btn_frame = ttk.Frame(self.schedule_frame)
        btn_frame.pack(pady=6)
        ttk.Button(btn_frame, text="Atualizar Cronograma", command=self.refresh_schedule_table).pack(side='left', padx=6)
        ttk.Button(btn_frame, text="Exportar CSV", command=self.export_schedule_csv).pack(side='left', padx=6)
        ttk.Button(btn_frame, text="Exportar XLSX", command=self.export_schedule_xlsx).pack(side='left', padx=6)

    def refresh_schedule_table(self):
        try:
            for iid in self.schedule_tree.get_children():
                self.schedule_tree.delete(iid)
            cursor = self.conn.cursor()
            cursor.execute("""
            SELECT p.nome, m.nome, t.nome, c.dia_semana, c.horario_inicio, c.horario_fim
            FROM cronogramas c
            JOIN professores p ON c.professor_id = p.id
            JOIN materias m ON c.materia_id = m.id
            JOIN turmas t ON c.turma_id = t.id
            ORDER BY CASE c.dia_semana
                WHEN 'Segunda' THEN 1
                WHEN 'Terça' THEN 2
                WHEN 'Quarta' THEN 3
                WHEN 'Quinta' THEN 4
                WHEN 'Sexta' THEN 5
                ELSE 99 END, c.horario_inicio, p.nome
            """)
            for prof, mat, turma, dia, ini, fim in cursor.fetchall():
                self.schedule_tree.insert('', 'end', values=(prof, mat, turma, dia, str(ini), str(fim)))
        except Exception as e:
            logging.error(f"Falha ao atualizar tabela de cronograma: {e}")

    def export_schedule_csv(self):
        # Exporta o cronograma atual para CSV simples no diretório atual
        try:
            import csv
            rows = []
            cursor = self.conn.cursor()
            cursor.execute("""
            SELECT p.nome, m.nome, t.nome, c.dia_semana, c.horario_inicio, c.horario_fim
            FROM cronogramas c
            JOIN professores p ON c.professor_id = p.id
            JOIN materias m ON c.materia_id = m.id
            JOIN turmas t ON c.turma_id = t.id
            ORDER BY CASE c.dia_semana
                WHEN 'Segunda' THEN 1
                WHEN 'Terça' THEN 2
                WHEN 'Quarta' THEN 3
                WHEN 'Quinta' THEN 4
                WHEN 'Sexta' THEN 5
                ELSE 99 END, c.horario_inicio, p.nome
            """)
            rows = cursor.fetchall()
            if not rows:
                show_message("Info", "Nenhum cronograma para exportar.", "info")
                return
            fname = f"cronograma_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            with open(fname, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Professor", "Matéria", "Turma", "Dia", "Início", "Fim"])
                for r in rows:
                    writer.writerow(r)
            show_message("Sucesso", f"Cronograma exportado para {fname}", "info")
        except Exception as e:
            logging.error(f"Erro ao exportar CSV: {e}")
            show_message("Erro", f"Falha ao exportar CSV: {e}", "error")

    def export_schedule_xlsx(self):
        try:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
            except Exception:
                show_message("Erro", "openpyxl não está instalado. Instale com: pip install openpyxl", "error")
                return

            cursor = self.conn.cursor()
            cursor.execute("""
            SELECT p.nome, m.nome, t.nome, c.dia_semana, c.horario_inicio, c.horario_fim
            FROM cronogramas c
            JOIN professores p ON c.professor_id = p.id
            JOIN materias m ON c.materia_id = m.id
            JOIN turmas t ON c.turma_id = t.id
            ORDER BY CASE c.dia_semana
                WHEN 'Segunda' THEN 1
                WHEN 'Terça' THEN 2
                WHEN 'Quarta' THEN 3
                WHEN 'Quinta' THEN 4
                WHEN 'Sexta' THEN 5
                ELSE 99 END, c.horario_inicio, p.nome
            """)
            rows = cursor.fetchall()
            if not rows:
                show_message("Info", "Nenhum cronograma para exportar.", "info")
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Cronograma"
            headers = ["Professor", "Matéria", "Turma", "Dia", "Início", "Fim"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill

            for r in rows:
                ws.append([r[0], r[1], r[2], r[3], str(r[4]), str(r[5])])

            # Ajusta largura das colunas
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)

            # Freeze header
            ws.freeze_panes = "A2"

            fname = f"cronograma_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(fname)
            show_message("Sucesso", f"Cronograma exportado para {fname}", "info")
        except Exception as e:
            logging.error(f"Erro ao exportar XLSX: {e}")
            show_message("Erro", f"Falha ao exportar XLSX: {e}", "error")

    def clear_prof(self):
        self.prof_nome.delete(0, tk.END)
        for var in self.prof_disp_vars.values():
            var.set(False)
        self.prof_pref_mat.delete(0, tk.END)
        self.prof_pref_dia.set("")

    def clear_mat(self):
        self.mat_nome.delete(0, tk.END)
        self.mat_carga.delete(0, tk.END)

    def clear_tur(self):
        self.tur_nome.delete(0, tk.END)
        self.tur_ano.delete(0, tk.END)

    def add_prof(self):
        nome = self.prof_nome.get().strip()
        disp = [dia for dia, var in self.prof_disp_vars.items() if var.get()]
        disp_str = ','.join(disp)
        pref_mat = self.prof_pref_mat.get().strip()
        pref_dia = self.prof_pref_dia.get().strip()
        pref = f"{pref_mat}:{pref_dia}" if pref_mat and pref_dia else ""
        if not nome:
            show_message("Erro", "O nome do professor é obrigatório!", "error")
            return
        try:
            cursor = self.conn.cursor()
            cursor.execute("INSERT INTO professores (nome, disponibilidade, preferencias) VALUES (%s, %s, %s)", (nome, disp_str, pref))
            self.conn.commit()
            show_message("Sucesso", f"Professor '{nome}' adicionado!", "info")
            self.clear_prof()
            self.refresh_prof_list()
        except Error as e:
            show_message("Erro", f"Falha ao adicionar professor: {e}", "error")

    def add_mat(self):
        nome = self.mat_nome.get().strip()
        carga = self.mat_carga.get().strip()
        if not nome:
            show_message("Erro", "O nome da matéria é obrigatório!", "error")
            return
        try:
            carga = int(carga)
            if carga <= 0:
                raise ValueError("Carga horária deve ser um número positivo!")
        except ValueError:
            show_message("Erro", "Carga horária deve ser um número inteiro positivo!", "error")
            return
        try:
            cursor = self.conn.cursor()
            cursor.execute("INSERT INTO materias (nome, carga_horaria) VALUES (%s, %s)", (nome, carga))
            self.conn.commit()
            show_message("Sucesso", f"Matéria '{nome}' adicionada!", "info")
            self.clear_mat()
            self.refresh_mat_list()
        except Error as e:
            show_message("Erro", f"Falha ao adicionar matéria: {e}", "error")

    def add_tur(self):
        nome = self.tur_nome.get().strip()
        ano = self.tur_ano.get().strip()
        if not nome:
            show_message("Erro", "O nome da turma é obrigatório!", "error")
            return
        try:
            ano = int(ano)
            if ano <= 0:
                raise ValueError("Ano deve ser um número positivo!")
        except ValueError:
            show_message("Erro", "Ano deve ser um número inteiro positivo!", "error")
            return
        try:
            cursor = self.conn.cursor()
            cursor.execute("INSERT INTO turmas (nome, ano) VALUES (%s, %s)", (nome, ano))
            self.conn.commit()
            show_message("Sucesso", f"Turma '{nome}' adicionada!", "info")
            self.clear_tur()
            self.refresh_tur_list()
        except Error as e:
            show_message("Erro", f"Falha ao adicionar turma: {e}", "error")

    def list_data(self):
        self.data_text.delete(1.0, tk.END)
        cursor = self.conn.cursor()

        self.data_text.insert(tk.END, "Professores:\n")
        cursor.execute("SELECT nome, disponibilidade, preferencias FROM professores")
        for nome, disp, pref in cursor.fetchall():
            self.data_text.insert(tk.END, f"- {nome} (Disp: {disp or 'Nenhuma'}, Pref: {pref or 'Nenhuma'})\n")

        self.data_text.insert(tk.END, "\nMatérias:\n")
        cursor.execute("SELECT nome, carga_horaria FROM materias")
        for nome, carga in cursor.fetchall():
            self.data_text.insert(tk.END, f"- {nome} (Carga: {carga}h)\n")

        self.data_text.insert(tk.END, "\nTurmas:\n")
        cursor.execute("SELECT nome, ano FROM turmas")
        for nome, ano in cursor.fetchall():
            self.data_text.insert(tk.END, f"- {nome} (Ano: {ano})\n")

    def generate(self):
        # Limpa apenas cronogramas antigos antes de gerar novo cronograma
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM cronogramas")
        self.conn.commit()
        ok = optimize_schedule(self.conn)
        if ok:
            # Atualiza a aba de tabela e seleciona-a
            self.refresh_schedule_table()
            try:
                self.notebook.select(self.schedule_frame)
            except Exception:
                pass
        else:
            # Se não foi possível gerar, atualiza área de texto para mensagem
            self.display_schedules()

    def display_schedules(self):
        self.text_area.delete(1.0, tk.END)
        cursor = self.conn.cursor()
        cursor.execute("""
        SELECT p.nome, m.nome, t.nome, c.dia_semana, c.horario_inicio, c.horario_fim
        FROM cronogramas c
        JOIN professores p ON c.professor_id = p.id
        JOIN materias m ON c.materia_id = m.id
        JOIN turmas t ON c.turma_id = t.id
        ORDER BY p.nome, c.dia_semana, c.horario_inicio
        """)
        results = cursor.fetchall()
        if not results:
            self.text_area.insert(tk.END, "Nenhum cronograma gerado ainda.\n")
            return
        current_prof = ""
        for row in results:
            prof, mat, tur, dia, ini, fim = row
            if prof != current_prof:
                self.text_area.insert(tk.END, f"\nCronograma para {prof}:\n")
                current_prof = prof
            self.text_area.insert(tk.END, f"- {dia}: {mat} para {tur} das {ini} às {fim}\n")

    # ----------------- CRUD helpers for Professores / Materias / Turmas -----------------
    def refresh_prof_list(self):
        try:
            for iid in self.prof_tree.get_children():
                self.prof_tree.delete(iid)
            cursor = self.conn.cursor()
            cursor.execute("SELECT id, nome, disponibilidade, preferencias FROM professores ORDER BY id")
            for pid, nome, disp, pref in cursor.fetchall():
                self.prof_tree.insert('', 'end', iid=str(pid), values=(pid, nome, disp or '', pref or ''))
        except Exception as e:
            logging.error(f"Falha ao atualizar lista de professores: {e}")

    def edit_prof(self):
        sel = self.prof_tree.selection()
        if not sel:
            show_message("Erro", "Nenhum professor selecionado para editar.", "error")
            return
        pid = int(sel[0])
        cursor = self.conn.cursor()
        cursor.execute("SELECT nome, disponibilidade, preferencias FROM professores WHERE id = %s", (pid,))
        row = cursor.fetchone()
        if not row:
            show_message("Erro", "Professor não encontrado.", "error")
            return
        nome, disp, pref = row

        win = tk.Toplevel(self.root)
        win.title("Editar Professor")

        ttk.Label(win, text="Nome:").grid(row=0, column=0, padx=8, pady=6, sticky='w')
        nome_e = tk.Entry(win, font=("Segoe UI", 11))
        nome_e.grid(row=0, column=1, padx=8, pady=6)
        nome_e.insert(0, nome)

        dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta']
        ttk.Label(win, text="Disponibilidade:").grid(row=1, column=0, padx=8, pady=6, sticky='w')
        disp_vars = {}
        dias_frame = ttk.Frame(win)
        dias_frame.grid(row=1, column=1, padx=0, pady=0, sticky='w')
        cur_disp = disp.split(',') if disp else []
        for i, d in enumerate(dias_semana):
            var = tk.BooleanVar(value=(d in cur_disp))
            cb = ttk.Checkbutton(dias_frame, text=d, variable=var)
            cb.pack(side='left', padx=2)
            disp_vars[d] = var

        ttk.Label(win, text="Preferência (Matéria):").grid(row=2, column=0, padx=8, pady=6, sticky='w')
        pref_mat_e = tk.Entry(win, font=("Segoe UI", 11))
        pref_mat_e.grid(row=2, column=1, padx=8, pady=6)
        pref_dia_e = ttk.Combobox(win, values=dias_semana, state='readonly')
        pref_dia_e.grid(row=3, column=1, padx=8, pady=6)
        if pref and ':' in pref:
            pm, pd = pref.split(':', 1)
            pref_mat_e.insert(0, pm)
            pref_dia_e.set(pd)

        def save():
            new_nome = nome_e.get().strip()
            new_disp = [d for d, v in disp_vars.items() if v.get()]
            new_pref_mat = pref_mat_e.get().strip()
            new_pref_dia = pref_dia_e.get().strip()
            new_pref = f"{new_pref_mat}:{new_pref_dia}" if new_pref_mat and new_pref_dia else ""
            try:
                cursor.execute("UPDATE professores SET nome=%s, disponibilidade=%s, preferencias=%s WHERE id=%s",
                               (new_nome, ','.join(new_disp), new_pref, pid))
                self.conn.commit()
                win.destroy()
                self.refresh_prof_list()
                show_message("Sucesso", "Professor atualizado.", "info")
            except Exception as e:
                show_message("Erro", f"Falha ao atualizar professor: {e}", "error")

        ttk.Button(win, text="Salvar", command=save).grid(row=4, column=0, pady=10)
        ttk.Button(win, text="Cancelar", command=win.destroy).grid(row=4, column=1, pady=10)

    def delete_prof(self):
        sel = self.prof_tree.selection()
        if not sel:
            show_message("Erro", "Nenhum professor selecionado para exclusão.", "error")
            return
        pid = int(sel[0])
        if not messagebox.askyesno("Confirmar", "Excluir o professor selecionado?"):
            return
        try:
            cursor = self.conn.cursor()
            cursor.execute("DELETE FROM professores WHERE id=%s", (pid,))
            self.conn.commit()
            self.refresh_prof_list()
            show_message("Sucesso", "Professor excluído.", "info")
        except Exception as e:
            show_message("Erro", f"Falha ao excluir professor: {e}", "error")

    def refresh_mat_list(self):
        try:
            for iid in self.mat_tree.get_children():
                self.mat_tree.delete(iid)
            cursor = self.conn.cursor()
            cursor.execute("SELECT id, nome, carga_horaria FROM materias ORDER BY id")
            for mid, nome, carga in cursor.fetchall():
                self.mat_tree.insert('', 'end', iid=str(mid), values=(mid, nome, carga))
        except Exception as e:
            logging.error(f"Falha ao atualizar lista de matérias: {e}")

    def edit_mat(self):
        sel = self.mat_tree.selection()
        if not sel:
            show_message("Erro", "Nenhuma matéria selecionada para editar.", "error")
            return
        mid = int(sel[0])
        cursor = self.conn.cursor()
        cursor.execute("SELECT nome, carga_horaria FROM materias WHERE id=%s", (mid,))
        row = cursor.fetchone()
        if not row:
            show_message("Erro", "Matéria não encontrada.", "error")
            return
        nome, carga = row

        win = tk.Toplevel(self.root)
        win.title("Editar Matéria")
        ttk.Label(win, text="Nome:").grid(row=0, column=0, padx=8, pady=6, sticky='w')
        nome_e = tk.Entry(win, font=("Segoe UI", 11))
        nome_e.grid(row=0, column=1, padx=8, pady=6)
        nome_e.insert(0, nome)
        ttk.Label(win, text="Carga Horária:").grid(row=1, column=0, padx=8, pady=6, sticky='w')
        carga_e = tk.Entry(win, font=("Segoe UI", 11))
        carga_e.grid(row=1, column=1, padx=8, pady=6)
        carga_e.insert(0, str(carga))

        def save():
            n = nome_e.get().strip()
            try:
                c = int(carga_e.get().strip())
                cursor.execute("UPDATE materias SET nome=%s, carga_horaria=%s WHERE id=%s", (n, c, mid))
                self.conn.commit()
                win.destroy()
                self.refresh_mat_list()
                show_message("Sucesso", "Matéria atualizada.", "info")
            except Exception as e:
                show_message("Erro", f"Falha ao atualizar matéria: {e}", "error")

        ttk.Button(win, text="Salvar", command=save).grid(row=2, column=0, pady=10)
        ttk.Button(win, text="Cancelar", command=win.destroy).grid(row=2, column=1, pady=10)

    def delete_mat(self):
        sel = self.mat_tree.selection()
        if not sel:
            show_message("Erro", "Nenhuma matéria selecionada para exclusão.", "error")
            return
        mid = int(sel[0])
        if not messagebox.askyesno("Confirmar", "Excluir a matéria selecionada?"):
            return
        try:
            cursor = self.conn.cursor()
            cursor.execute("DELETE FROM materias WHERE id=%s", (mid,))
            self.conn.commit()
            self.refresh_mat_list()
            show_message("Sucesso", "Matéria excluída.", "info")
        except Exception as e:
            show_message("Erro", f"Falha ao excluir matéria: {e}", "error")

    def refresh_tur_list(self):
        try:
            for iid in self.tur_tree.get_children():
                self.tur_tree.delete(iid)
            cursor = self.conn.cursor()
            cursor.execute("SELECT id, nome, ano FROM turmas ORDER BY id")
            for tid, nome, ano in cursor.fetchall():
                self.tur_tree.insert('', 'end', iid=str(tid), values=(tid, nome, ano))
        except Exception as e:
            logging.error(f"Falha ao atualizar lista de turmas: {e}")

    def edit_tur(self):
        sel = self.tur_tree.selection()
        if not sel:
            show_message("Erro", "Nenhuma turma selecionada para editar.", "error")
            return
        tid = int(sel[0])
        cursor = self.conn.cursor()
        cursor.execute("SELECT nome, ano FROM turmas WHERE id=%s", (tid,))
        row = cursor.fetchone()
        if not row:
            show_message("Erro", "Turma não encontrada.", "error")
            return
        nome, ano = row

        win = tk.Toplevel(self.root)
        win.title("Editar Turma")
        ttk.Label(win, text="Nome:").grid(row=0, column=0, padx=8, pady=6, sticky='w')
        nome_e = tk.Entry(win, font=("Segoe UI", 11))
        nome_e.grid(row=0, column=1, padx=8, pady=6)
        nome_e.insert(0, nome)
        ttk.Label(win, text="Ano:").grid(row=1, column=0, padx=8, pady=6, sticky='w')
        ano_e = tk.Entry(win, font=("Segoe UI", 11))
        ano_e.grid(row=1, column=1, padx=8, pady=6)
        ano_e.insert(0, str(ano))

        def save():
            n = nome_e.get().strip()
            try:
                a = int(ano_e.get().strip())
                cursor.execute("UPDATE turmas SET nome=%s, ano=%s WHERE id=%s", (n, a, tid))
                self.conn.commit()
                win.destroy()
                self.refresh_tur_list()
                show_message("Sucesso", "Turma atualizada.", "info")
            except Exception as e:
                show_message("Erro", f"Falha ao atualizar turma: {e}", "error")

        ttk.Button(win, text="Salvar", command=save).grid(row=2, column=0, pady=10)
        ttk.Button(win, text="Cancelar", command=win.destroy).grid(row=2, column=1, pady=10)

    def delete_tur(self):
        sel = self.tur_tree.selection()
        if not sel:
            show_message("Erro", "Nenhuma turma selecionada para exclusão.", "error")
            return
        tid = int(sel[0])
        if not messagebox.askyesno("Confirmar", "Excluir a turma selecionada?"):
            return
        try:
            cursor = self.conn.cursor()
            cursor.execute("DELETE FROM turmas WHERE id=%s", (tid,))
            self.conn.commit()
            self.refresh_tur_list()
            show_message("Sucesso", "Turma excluída.", "info")
        except Exception as e:
            show_message("Erro", f"Falha ao excluir turma: {e}", "error")


def print_schedule_console(connection):
    cursor = connection.cursor()
    cursor.execute("""
    SELECT p.nome, m.nome, t.nome, c.dia_semana, c.horario_inicio, c.horario_fim
    FROM cronogramas c
    JOIN professores p ON c.professor_id = p.id
    JOIN materias m ON c.materia_id = m.id
    JOIN turmas t ON c.turma_id = t.id
    ORDER BY p.nome, c.dia_semana, c.horario_inicio
    """)
    rows = cursor.fetchall()
    if not rows:
        print("Nenhum cronograma gerado ainda.")
        return
    current = None
    for prof, mat, tur, dia, ini, fim in rows:
        if prof != current:
            print(f"\nCronograma para {prof}:")
            current = prof
        print(f"- {dia}: {mat} para {tur} das {ini} às {fim}")


def seed_sample_data(conn):
    cursor = conn.cursor()
    # minimal check to avoid duplicate seeding
    cursor.execute("SELECT COUNT(*) FROM professores")
    if cursor.fetchone()[0] > 0:
        logging.info("Parece que já existem professores cadastrados; seed ignorado.")
        return
    # Professores
    cursor.execute("INSERT INTO professores (nome, disponibilidade, preferencias) VALUES (%s, %s, %s)", ("Robson", "Segunda,Terça,Quarta", "Devops:Quarta"))
    cursor.execute("INSERT INTO professores (nome, disponibilidade, preferencias) VALUES (%s, %s, %s)", ("Belloni", "Segunda,Terça,Quarta", "TikTok:Terça"))
    # Materias
    cursor.execute("INSERT INTO materias (nome, carga_horaria) VALUES (%s, %s)", ("Devops", 10))
    cursor.execute("INSERT INTO materias (nome, carga_horaria) VALUES (%s, %s)", ("TikTok", 10))
    # Turmas
    cursor.execute("INSERT INTO turmas (nome, ano) VALUES (%s, %s)", ("Oitavo", 2))
    conn.commit()
    logging.info("Dados de exemplo inseridos (seed).")


def main():
    parser = argparse.ArgumentParser(description='School Scheduler')
    parser.add_argument('--headless', action='store_true', help='Run in headless mode (no GUI)')
    parser.add_argument('--seed-sample', action='store_true', help='Seed sample data before running (headless recommended)')
    args = parser.parse_args()

    conn = create_connection()
    if not conn:
        logging.error("Não foi possível conectar ao banco de dados. Saindo.")
        sys.exit(1)
    create_tables(conn)

    if args.headless:
        if args.seed_sample:
            seed_sample_data(conn)
        # ensure cronogramas is clean
        cur = conn.cursor()
        cur.execute("DELETE FROM cronogramas")
        conn.commit()
        ok = optimize_schedule(conn)
        if ok:
            print_schedule_console(conn)
        conn.close()
        return

    # GUI path
    try:
        root = tk.Tk()
        SchoolApp(root, conn)
        root.mainloop()
    except _tkinter.TclError as e:
        logging.error(f"Erro ao iniciar GUI: {e}")
        print("Erro ao iniciar interface gráfica (verifique DISPLAY). Tente rodar com --headless.")
    finally:
        conn.close()


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s')
    main()
